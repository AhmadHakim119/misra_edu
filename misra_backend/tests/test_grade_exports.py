import io
import os
import unittest

from openpyxl import load_workbook
from sqlalchemy import create_engine
from sqlalchemy.orm import sessionmaker


os.environ.setdefault("GEMINI_API_KEY", "test-key-not-used")

from database import Base  # noqa: E402
import models  # noqa: E402,F401
from models import Answer, Course, Exam, Institution, Question, Student, Submission, User  # noqa: E402
from services.grade_export_service import (  # noqa: E402
    blackboard_omission_reasons,
    build_csv,
    build_grade_export,
    build_export_preflight,
    build_xlsx,
)


class GradeExportTests(unittest.TestCase):
    def setUp(self):
        engine = create_engine("sqlite+pysqlite:///:memory:")
        Base.metadata.create_all(engine)
        self.db = sessionmaker(bind=engine)()
        institution = Institution(id="institution-1", name="Test University")
        teacher = User(
            id="teacher-1",
            institution_id=institution.id,
            email="teacher@example.edu",
            hashed_password="not-used",
            role="teacher",
        )
        course = Course(
            id="course-1",
            institution_id=institution.id,
            teacher_id=teacher.id,
            course_code="CS2071",
            title="Database Systems",
        )
        exam = Exam(
            id="exam-1",
            institution_id=institution.id,
            course_id=course.id,
            title="Midterm",
            language="en",
        )
        question = Question(
            id="question-1",
            institution_id=institution.id,
            exam_id=exam.id,
            question_number="1",
            question_text="Create an EER diagram.",
            max_score=10,
            rubric_json={},
            order_index=1,
            language="en",
        )
        student = Student(
            id="student-1",
            institution_id=institution.id,
            student_number="S1001",
            full_name="Student One",
            email="student@example.edu",
        )
        submission = Submission(
            id="submission-1",
            institution_id=institution.id,
            exam_id=exam.id,
            student_id=student.id,
            original_file_path="unused.pdf",
            status="reviewed",
        )
        answer = Answer(
            id="answer-1",
            institution_id=institution.id,
            submission_id=submission.id,
            question_id=question.id,
            score=8,
            max_score=10,
            needs_review=False,
            review_status="approved",
        )
        self.db.add_all([institution, teacher, course, exam, question, student, submission, answer])
        self.db.commit()

    def tearDown(self):
        self.db.close()

    def test_blackboard_profile_contains_ready_grade(self):
        export = build_grade_export("exam-1", "institution-1", self.db)
        content, omitted = build_csv(export, "blackboard", "student_number")
        text = content.decode("utf-8-sig")

        self.assertEqual(omitted, 0)
        self.assertIn("Username", text)
        self.assertIn("S1001", text)
        self.assertIn(",8", text)

    def test_xlsx_has_gradebook_and_question_breakdown(self):
        export = build_grade_export("exam-1", "institution-1", self.db)
        workbook = load_workbook(io.BytesIO(build_xlsx(export)), data_only=False)

        self.assertEqual(workbook.sheetnames, ["Gradebook", "Question breakdown", "Export notes"])
        self.assertEqual(workbook["Gradebook"]["A2"].value, "S1001")
        self.assertEqual(workbook["Gradebook"]["D2"].value, 8)
        self.assertEqual(workbook["Question breakdown"]["C2"].value, "1")

    def test_cross_institution_export_is_rejected(self):
        with self.assertRaises(ValueError):
            build_grade_export("exam-1", "different-institution", self.db)

    def test_blackboard_omission_explains_missing_identifier(self):
        submission = self.db.query(Submission).filter(Submission.id == "submission-1").one()
        submission.student_id = None
        submission.extracted_student_name = "Student One"
        self.db.commit()

        export = build_grade_export("exam-1", "institution-1", self.db)
        _, omitted = build_csv(export, "blackboard", "student_number")
        reasons = blackboard_omission_reasons(export, "student_number")

        self.assertEqual(omitted, 1)
        self.assertEqual(reasons["missing_identifier"], 1)
        self.assertEqual(reasons["incomplete_grading"], 0)

    def test_blackboard_preflight_keeps_identity_context_out_of_import_columns(self):
        export = build_grade_export("exam-1", "institution-1", self.db)
        preflight = build_export_preflight(export, "student_number")

        self.assertEqual(preflight["counts"]["ready"], 1)
        self.assertEqual(preflight["rows"][0]["student_name"], "Student One")
        self.assertEqual(preflight["rows"][0]["username"], "S1001")
        self.assertEqual(preflight["rows"][0]["issues"], [])


if __name__ == "__main__":
    unittest.main()
