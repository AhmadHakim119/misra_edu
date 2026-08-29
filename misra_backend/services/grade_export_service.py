from __future__ import annotations

import csv
import io
import re
from collections import defaultdict
from dataclasses import dataclass
from decimal import Decimal

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from sqlalchemy.orm import Session

from models import Answer, Course, Exam, Question, Student, Submission


FORMULA_PREFIXES = ("=", "+", "-", "@")


@dataclass(frozen=True)
class GradeExport:
    exam: Exam
    course: Course | None
    rows: list[dict]
    question_rows: list[dict]


def _safe_text(value: object | None) -> str:
    text = "" if value is None else str(value)
    return f"'{text}" if text.startswith(FORMULA_PREFIXES) else text


def _number(value: object | None) -> float | None:
    if value is None:
        return None
    if isinstance(value, Decimal):
        return float(value)
    return float(value)


def _effective_score(answer: Answer | None) -> float | None:
    if not answer:
        return None
    return _number(answer.teacher_override_score if answer.teacher_override_score is not None else answer.score)


def build_grade_export(exam_id: str, institution_id: str, db: Session) -> GradeExport:
    exam = db.query(Exam).filter(Exam.id == exam_id, Exam.institution_id == institution_id).first()
    if not exam:
        raise ValueError("Assessment not found")

    course = db.query(Course).filter(Course.id == exam.course_id).first()
    questions = (
        db.query(Question)
        .filter(Question.exam_id == exam.id)
        .order_by(Question.order_index.asc(), Question.question_number.asc())
        .all()
    )
    submissions = (
        db.query(Submission)
        .filter(Submission.exam_id == exam.id)
        .order_by(Submission.uploaded_at.asc())
        .all()
    )
    submission_ids = [submission.id for submission in submissions]
    answers = (
        db.query(Answer).filter(Answer.submission_id.in_(submission_ids)).all()
        if submission_ids
        else []
    )
    answers_by_submission: dict[str, dict[str, Answer]] = defaultdict(dict)
    for answer in answers:
        answers_by_submission[answer.submission_id][answer.question_id] = answer

    student_ids = [submission.student_id for submission in submissions if submission.student_id]
    students = db.query(Student).filter(Student.id.in_(student_ids)).all() if student_ids else []
    students_by_id = {student.id: student for student in students}
    total_max = sum(float(question.max_score) for question in questions)

    rows: list[dict] = []
    question_rows: list[dict] = []
    for submission in submissions:
        student = students_by_id.get(submission.student_id)
        submission_answers = answers_by_submission.get(submission.id, {})
        scores = [_effective_score(submission_answers.get(question.id)) for question in questions]
        graded_count = sum(score is not None for score in scores)
        complete = bool(questions) and graded_count == len(questions)
        needs_review = any(
            answer.needs_review or answer.review_status == "pending"
            for answer in submission_answers.values()
        )
        score = sum(score or 0 for score in scores) if complete else None
        student_number = student.student_number if student else submission.extracted_student_number
        student_name = student.full_name if student else submission.extracted_student_name
        email = student.email if student else None
        rows.append(
            {
                "submission_id": submission.id,
                "student_number": student_number,
                "student_name": student_name,
                "email": email,
                "score": score,
                "max_score": total_max,
                "percentage": (score / total_max * 100) if score is not None and total_max else None,
                "graded_questions": graded_count,
                "question_count": len(questions),
                "needs_review": needs_review,
                "ready_for_lms": complete and not needs_review,
                "submission_status": submission.status,
            }
        )
        for question, question_score in zip(questions, scores):
            answer = submission_answers.get(question.id)
            question_rows.append(
                {
                    "submission_id": submission.id,
                    "student_number": student_number,
                    "student_name": student_name,
                    "question_number": question.question_number,
                    "score": question_score,
                    "max_score": float(question.max_score),
                    "review_status": answer.review_status if answer else "not_graded",
                }
            )

    return GradeExport(exam=exam, course=course, rows=rows, question_rows=question_rows)


def _grade_column_title(export: GradeExport) -> str:
    title = export.exam.title.strip()
    return f"{title} [Total Pts: {sum(row['max_score'] for row in export.rows[:1]):g}]"


def build_csv(export: GradeExport, profile: str, identifier: str) -> tuple[bytes, int]:
    buffer = io.StringIO(newline="")
    writer = csv.writer(buffer, lineterminator="\r\n")
    omitted = 0
    if profile == "blackboard":
        writer.writerow(["Username", _grade_column_title(export)])
        for row in export.rows:
            username = row["email"] if identifier == "email" else row["student_number"]
            if not username or not row["ready_for_lms"]:
                omitted += 1
                continue
            writer.writerow([_safe_text(username), row["score"]])
    else:
        writer.writerow(
            [
                "Student ID",
                "Student Name",
                "Email",
                "Score",
                "Max Score",
                "Percentage",
                "Ready for LMS",
                "Review Status",
                "Submission ID",
            ]
        )
        for row in export.rows:
            writer.writerow(
                [
                    _safe_text(row["student_number"]),
                    _safe_text(row["student_name"]),
                    _safe_text(row["email"]),
                    row["score"] if row["score"] is not None else "",
                    row["max_score"],
                    row["percentage"] if row["percentage"] is not None else "",
                    "Yes" if row["ready_for_lms"] else "No",
                    "Needs review" if row["needs_review"] else "Clear",
                    row["submission_id"],
                ]
            )
    return ("\ufeff" + buffer.getvalue()).encode("utf-8"), omitted


def blackboard_omission_reasons(export: GradeExport, identifier: str) -> dict[str, int]:
    """Explain why rows cannot be safely imported into Blackboard.

    Blackboard matches grades to its roster by Username. A student name is not
    a safe substitute, so missing identifiers remain an explicit blocker.
    """
    reasons = {
        "missing_identifier": 0,
        "incomplete_grading": 0,
        "needs_review": 0,
    }
    for row in export.rows:
        username = row["email"] if identifier == "email" else row["student_number"]
        if not username:
            reasons["missing_identifier"] += 1
        if row["graded_questions"] != row["question_count"]:
            reasons["incomplete_grading"] += 1
        if row["needs_review"]:
            reasons["needs_review"] += 1
    return reasons


def build_export_preflight(export: GradeExport, identifier: str) -> dict:
    rows = []
    counts = {
        "ready": 0,
        "missing_name": 0,
        "missing_identifier": 0,
        "incomplete_grading": 0,
        "needs_review": 0,
    }
    for row in export.rows:
        username = row["email"] if identifier == "email" else row["student_number"]
        issues = []
        if not row["student_name"]:
            issues.append({"code": "missing_name", "message": "Student name is missing", "blocking": False})
            counts["missing_name"] += 1
        if not username:
            label = "Blackboard email" if identifier == "email" else "Blackboard username / student number"
            issues.append({"code": "missing_identifier", "message": f"{label} is missing", "blocking": True})
            counts["missing_identifier"] += 1
        if row["graded_questions"] != row["question_count"]:
            issues.append({"code": "incomplete_grading", "message": "Grading is incomplete", "blocking": True})
            counts["incomplete_grading"] += 1
        if row["needs_review"]:
            issues.append({"code": "needs_review", "message": "Instructor review is pending", "blocking": True})
            counts["needs_review"] += 1
        ready = not any(issue["blocking"] for issue in issues)
        if ready:
            counts["ready"] += 1
        rows.append(
            {
                "submission_id": row["submission_id"],
                "student_name": row["student_name"],
                "username": username,
                "score": row["score"],
                "max_score": row["max_score"],
                "ready": ready,
                "issues": issues,
            }
        )
    return {
        "assessment": export.exam.title,
        "identifier": identifier,
        "grade_column": _grade_column_title(export),
        "counts": counts,
        "rows": rows,
    }


def _style_sheet(sheet, widths: list[float]) -> None:
    header_fill = PatternFill("solid", fgColor="DDE9DF")
    for cell in sheet[1]:
        cell.font = Font(bold=True, color="294536")
        cell.fill = header_fill
        cell.alignment = Alignment(vertical="center")
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    for index, width in enumerate(widths, start=1):
        sheet.column_dimensions[get_column_letter(index)].width = width


def build_xlsx(export: GradeExport) -> bytes:
    workbook = Workbook()
    gradebook = workbook.active
    gradebook.title = "Gradebook"
    gradebook.append(
        [
            "Student ID",
            "Student Name",
            "Email",
            "Score",
            "Max Score",
            "Percentage",
            "Ready for LMS",
            "Review Status",
            "Submission ID",
        ]
    )
    for row in export.rows:
        gradebook.append(
            [
                _safe_text(row["student_number"]),
                _safe_text(row["student_name"]),
                _safe_text(row["email"]),
                row["score"],
                row["max_score"],
                row["percentage"] / 100 if row["percentage"] is not None else None,
                "Yes" if row["ready_for_lms"] else "No",
                "Needs review" if row["needs_review"] else "Clear",
                row["submission_id"],
            ]
        )
    for cell in gradebook["F"][1:]:
        cell.number_format = "0.00%"
    _style_sheet(gradebook, [18, 26, 30, 12, 12, 14, 16, 16, 38])

    breakdown = workbook.create_sheet("Question breakdown")
    breakdown.append(
        ["Student ID", "Student Name", "Question", "Score", "Max Score", "Review Status", "Submission ID"]
    )
    for row in export.question_rows:
        breakdown.append(
            [
                _safe_text(row["student_number"]),
                _safe_text(row["student_name"]),
                _safe_text(row["question_number"]),
                row["score"],
                row["max_score"],
                row["review_status"],
                row["submission_id"],
            ]
        )
    _style_sheet(breakdown, [18, 26, 14, 12, 12, 18, 38])

    metadata = workbook.create_sheet("Export notes")
    metadata.append(["Field", "Value"])
    metadata.append(["Assessment", export.exam.title])
    metadata.append(["Course", export.course.title if export.course else ""])
    metadata.append(["Course code", export.course.course_code if export.course else ""])
    metadata.append(["Term", export.course.term if export.course else ""])
    metadata.append(["Rule", "Teacher overrides are exported when present; unresolved or incomplete rows are not LMS-ready."])
    _style_sheet(metadata, [20, 90])
    metadata.column_dimensions["B"].width = 90
    for cell in metadata["B"]:
        cell.alignment = Alignment(wrap_text=True, vertical="top")

    output = io.BytesIO()
    workbook.save(output)
    return output.getvalue()


def safe_filename(value: str) -> str:
    slug = re.sub(r"[^A-Za-z0-9._-]+", "-", value.strip()).strip("-.")
    return slug or "assessment"
